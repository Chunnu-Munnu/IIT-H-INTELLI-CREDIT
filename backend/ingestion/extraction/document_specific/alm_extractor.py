"""
ALM (Asset-Liability Management) Extractor
Handles PDF and XLSX formats showing maturity bucket tables.
Extracts gap analysis, liquidity risk signals, rate sensitivity.
"""
import re
from loguru import logger

NUMBER_RE = re.compile(r'[-−]?\s*[\d,]+(?:\.\d+)?')

TIME_BUCKETS = [
    ("1_day", ["1 day", "one day", "next day", "overnight"]),
    ("2_7_days", ["2-7 days", "2 to 7", "week", "one week"]),
    ("8_14_days", ["8-14 days", "8 to 14", "fortnightly", "14 days"]),
    ("15_28_days", ["15-28 days", "15 to 28", "one month", "28 days"]),
    ("3_months", ["1-3 months", "1 to 3 month", "3 months", "quarterly"]),
    ("6_months", ["3-6 months", "3 to 6", "6 months", "half year"]),
    ("1_year", ["6-12 months", "6 to 12", "1 year", "one year"]),
    ("3_years", ["1-3 years", "1 to 3 year", "3 years"]),
    ("5_years", ["3-5 years", "3 to 5", "5 years", "medium term"]),
    ("over_5_years", ["5+ years", "above 5", "over 5", "long term", "more than 5"]),
]


class ALMExtractor:

    def extract(self, file_path: str) -> dict:
        result = self._empty_result()
        ext = file_path.lower().split(".")[-1]
        try:
            if ext in ("xlsx", "xls"):
                return self._extract_from_excel(file_path, result)
            else:
                return self._extract_from_pdf(file_path, result)
        except Exception as e:
            logger.warning(f"[ALMExtractor] Failed: {e}")
            return result

    # ── EXCEL ────────────────────────────────────────────────────────────────

    def _extract_from_excel(self, file_path: str, result: dict) -> dict:
        rows = []
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                for row in ws.iter_rows(values_only=True):
                    rows.append([str(c) if c is not None else "" for c in row])
            wb.close()
        except Exception:
            try:
                import xlrd
                wb = xlrd.open_workbook(file_path)
                for sh in wb.sheets():
                    for r in range(sh.nrows):
                        rows.append([str(sh.cell_value(r, c)) for c in range(sh.ncols)])
            except Exception as e2:
                logger.warning(f"ALM Excel open failed: {e2}")
                return result

        self._extract_kpis("\n".join([" ".join(r) for r in rows]), result)
        return self._parse_table_rows(rows, result)

    # ── PDF ──────────────────────────────────────────────────────────────────

    def _extract_from_pdf(self, file_path: str, result: dict) -> dict:
        # Try pdfplumber for table extraction first
        try:
            import pdfplumber
            with pdfplumber.open(file_path) as pdf:
                for page in pdf.pages:
                    tables = page.extract_tables()
                    for table in tables:
                        if table and len(table) > 2:
                            r = self._parse_table_rows(table, result)
                            if r["buckets"]:
                                return r
        except Exception as e:
            logger.debug(f"pdfplumber ALM: {e}")

        # Fallback: raw text
        text = ""
        try:
            import fitz
            doc = fitz.open(file_path)
            for page in doc:
                text += page.get_text() + "\n"
            doc.close()
        except Exception:
            pass

        return self._parse_text_fallback(text.lower(), result)

    # ── KPI EXTRACTION ───────────────────────────────────────────────────────

    def _extract_kpis(self, text: str, result: dict):
        lcr_match = re.search(r'liquidity coverage ratio\s*[:\-]?\s*(\d+(?:\.\d+)?)%?', text.lower())
        if lcr_match:
            result["kpis"]["lcr"] = float(lcr_match.group(1))
        
        nsfr_match = re.search(r'net stable funding ratio\s*[:\-]?\s*(\d+(?:\.\d+)?)%?', text.lower())
        if nsfr_match:
            result["kpis"]["nsfr"] = float(nsfr_match.group(1))

    # ── PARSER ───────────────────────────────────────────────────────────────

    def _parse_table_rows(self, rows: list, result: dict) -> dict:
        """Try to find asset/liability rows and match them to time buckets."""
        asset_row = None
        liability_row = None

        for row in rows:
            row_lower = [str(c).lower().strip() for c in row]
            label = row_lower[0] if row_lower else ""

            # Identify asset row
            if any(kw in label for kw in ["total asset", "assets", "rate sensitive assets", "rsa"]):
                asset_row = row
            # Identify liability row
            elif any(kw in label for kw in ["total liabilit", "liabilities", "rate sensitive liabilit", "rsl"]):
                liability_row = row

        if asset_row and liability_row:
            self._map_row_to_buckets(asset_row, liability_row, result)
        elif rows:
            # Try to parse header-based matching
            self._parse_header_based(rows, result)

        self._compute_gaps(result)
        self._generate_risk_signals(result)
        logger.success(f"[ALMExtractor] Buckets extracted: {len(result['buckets'])}")
        return result

    def _map_row_to_buckets(self, asset_row: list, liability_row: list, result: dict):
        """Map column values to time buckets using column position."""
        # Skip the label column (col 0)
        asset_vals = self._extract_numbers_from_row(asset_row[1:])
        liab_vals = self._extract_numbers_from_row(liability_row[1:])

        bucket_keys = [b[0] for b in TIME_BUCKETS]
        for i, bkey in enumerate(bucket_keys):
            a = asset_vals[i] if i < len(asset_vals) else None
            l = liab_vals[i] if i < len(liab_vals) else None
            if a is not None or l is not None:
                result["buckets"][bkey] = {
                    "label": TIME_BUCKETS[i][1][0],
                    "assets_paise": int(round(a * 100)) if a else 0,
                    "liabilities_paise": int(round(l * 100)) if l else 0,
                    "gap_paise": 0,
                }

    def _parse_header_based(self, rows: list, result: dict):
        """Find bucket columns from header row then read data rows."""
        if not rows:
            return
        header = [str(c).lower() for c in rows[0]]
        bucket_col_map = {}  # bucket_key -> column index
        for col_idx, cell in enumerate(header):
            for bkey, keywords in TIME_BUCKETS:
                if any(kw in cell for kw in keywords):
                    bucket_col_map[bkey] = col_idx
                    break

        for row in rows[1:]:
            label = str(row[0]).lower() if row else ""
            is_asset = any(k in label for k in ["asset", "rsa"])
            is_liab = any(k in label for k in ["liab", "rsl"])
            if not (is_asset or is_liab):
                continue
            for bkey, col_idx in bucket_col_map.items():
                if col_idx < len(row):
                    val = self._parse_single_number(str(row[col_idx]))
                    if val is not None:
                        if bkey not in result["buckets"]:
                            result["buckets"][bkey] = {
                                "label": bkey.replace("_", " "),
                                "assets_paise": 0, "liabilities_paise": 0, "gap_paise": 0
                            }
                        if is_asset:
                            result["buckets"][bkey]["assets_paise"] = int(val * 100)
                        else:
                            result["buckets"][bkey]["liabilities_paise"] = int(val * 100)

    def _parse_text_fallback(self, text: str, result: dict) -> dict:
        """When tables fail, try regex approach on raw text."""
        for bkey, keywords in TIME_BUCKETS:
            for kw in keywords:
                idx = text.find(kw)
                if idx >= 0:
                    snippet = text[idx:idx+200]
                    nums = self._extract_numbers_from_row(snippet.split())
                    if len(nums) >= 2:
                        result["buckets"][bkey] = {
                            "label": kw,
                            "assets_paise": int(nums[0] * 100),
                            "liabilities_paise": int(nums[1] * 100),
                            "gap_paise": 0,
                        }
                        break
        self._compute_gaps(result)
        self._generate_risk_signals(result)
        return result

    def _compute_gaps(self, result: dict):
        total_assets = 0
        total_liab = 0
        for bkey, bdata in result["buckets"].items():
            gap = bdata["assets_paise"] - bdata["liabilities_paise"]
            bdata["gap_paise"] = gap
            total_assets += bdata["assets_paise"]
            total_liab += bdata["liabilities_paise"]
        result["total_assets_paise"] = total_assets
        result["total_liabilities_paise"] = total_liab
        result["overall_gap_paise"] = total_assets - total_liab

    def _generate_risk_signals(self, result: dict):
        signals = []
        # Short-term buckets = 1 day + 2-7 days + 8-14 days + 15-30 days
        short_term_buckets = ["1_day", "2_7_days", "8_14_days", "15_30_days"]
        short_gap = sum(result["buckets"].get(bk, {}).get("gap_paise", 0) for bk in short_term_buckets)

        CRORE_PAISE = 10_000_000 * 100
        if short_gap < -5 * CRORE_PAISE:
            signals.append({
                "signal_type": "ALM_SHORT_TERM_MISMATCH",
                "severity": "HIGH" if short_gap > -20 * CRORE_PAISE else "CRITICAL",
                "description": f"Short-term ALM gap (1 day to 30 days) is negative ₹{abs(short_gap) // CRORE_PAISE:.1f} Cr — liabilities exceed assets in near-term, indicating liquidity risk.",
                "five_c_mapping": "Capacity",
            })

        # Structural gap
        if result["overall_gap_paise"] < -10 * CRORE_PAISE:
            signals.append({
                "signal_type": "ALM_STRUCTURAL_DEFICIT",
                "severity": "MEDIUM",
                "description": f"Overall asset-liability gap is negative ₹{abs(result['overall_gap_paise']) // CRORE_PAISE:.1f} Cr — structural funding gap exists.",
                "five_c_mapping": "Capital",
            })

        result["risk_signals"] = signals

    # ── Helpers ──────────────────────────────────────────────────────────────

    def _extract_numbers_from_row(self, cells: list) -> list:
        numbers = []
        for cell in cells:
            n = self._parse_single_number(str(cell))
            if n is not None:
                numbers.append(n)
        return numbers

    def _parse_single_number(self, text: str) -> float | None:
        text = text.strip().replace(",", "").replace("(", "-").replace(")", "")
        try:
            return float(text)
        except ValueError:
            return None

    def _empty_result(self) -> dict:
        return {
            "buckets": {},
            "kpis": {"lcr": None, "nsfr": None},
            "total_assets_paise": 0,
            "total_liabilities_paise": 0,
            "overall_gap_paise": 0,
            "risk_signals": [],
            "extraction_source": "alm_extractor",
        }
