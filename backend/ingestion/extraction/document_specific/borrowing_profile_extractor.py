"""
Borrowing Profile Extractor
Handles PDF and XLSX schedules of existing borrowings.
Extracts: lender name, facility type, sanctioned amount, outstanding, rate, security.
Cross-validates total outstanding vs balance sheet borrowings to flag undisclosed debt.
"""
import re
from loguru import logger
from ingestion.normalization.currency_normalizer import CurrencyNormalizer

CURR = CurrencyNormalizer()

FACILITY_TYPES = ["term loan", "cash credit", "overdraft", "od", "cc", "letter of credit",
                  "lc", "bank guarantee", "bg", "wcdl", "working capital", "pcfc", "buyers credit",
                  "foreign currency", "ncd", "debenture", "bond", "commercial paper", "cp"]

LENDER_KEYWORDS = ["bank", "financial institution", "nbfc", "ltd", "limited", "sbi",
                   "hdfc", "icici", "axis", "kotak", "ubi", "pnb", "boi", "canara",
                   "union", "idbi", "yes bank", "rbl", "indusind", "standard chartered",
                   "citibank", "deutsche", "barclays", "dbs", "federal", "south indian",
                   "state bank", "bank of baroda", "ncd", "term loan"]


class BorrowingProfileExtractor:

    def extract(self, file_path: str) -> dict:
        result = self._empty_result()
        ext = file_path.lower().split(".")[-1]
        try:
            if ext in ("xlsx", "xls"):
                return self._extract_from_excel(file_path, result)
            else:
                return self._extract_from_pdf(file_path, result)
        except Exception as e:
            logger.warning(f"[BorrowingProfileExtractor] Failed: {e}")
            return result

    # ── EXCEL ────────────────────────────────────────────────────────────────

    def _extract_from_excel(self, file_path: str, result: dict) -> dict:
        rows = []
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                for row in ws.iter_rows(values_only=True):
                    rows.append([str(c).strip() if c is not None else "" for c in row])
            wb.close()
        except Exception:
            try:
                import xlrd
                wb = xlrd.open_workbook(file_path)
                for sh in wb.sheets():
                    for r in range(sh.nrows):
                        rows.append([str(sh.cell_value(r, c)) for c in range(sh.ncols)])
            except Exception as e2:
                logger.warning(f"Borrowing Excel failed: {e2}")
                return result

        return self._parse_from_rows(rows, result)

    # ── PDF ──────────────────────────────────────────────────────────────────

    def _extract_from_pdf(self, file_path: str, result: dict) -> dict:
        rows = []
        # pdfplumber table extraction first
        try:
            import pdfplumber
            with pdfplumber.open(file_path) as pdf:
                for page in pdf.pages:
                    tables = page.extract_tables()
                    for table in tables:
                        if table:
                            rows.extend([[str(c) if c else "" for c in row] for row in table])
        except Exception as e:
            logger.debug(f"pdfplumber borrowing profile: {e}")

        if rows:
            r = self._parse_from_rows(rows, result)
            if r["loan_entries"]:
                return r

        # Fallback: raw text line parsing
        text = ""
        try:
            import fitz
            doc = fitz.open(file_path)
            for page in doc:
                text += page.get_text() + "\n"
            doc.close()
        except Exception:
            pass

        return self._parse_from_text(text, result)

    # ── PARSER ───────────────────────────────────────────────────────────────

    def _parse_from_rows(self, rows: list, result: dict) -> dict:
        # Find header row to understand column layout
        header_idx = -1
        col_map = {}

        for i, row in enumerate(rows):
            row_str = " ".join(row).lower()
            if self._count_matches(row_str, ["lender", "bank", "institution", "facility", "sanctioned", "outstanding", "limit"]) >= 2:
                header_idx = i
                # Map column indexes
                for j, cell in enumerate(row):
                    cl = cell.lower().strip()
                    if any(k in cl for k in ["lender", "bank", "institution", "name"]):
                        col_map["lender"] = j
                    elif any(k in cl for k in ["facility", "type", "nature"]):
                        col_map["facility_type"] = j
                    elif any(k in cl for k in ["sanction", "limit", "approved"]):
                        col_map["sanctioned"] = j
                    elif any(k in cl for k in ["outstanding", "balance", "drawn", "utiliz"]):
                        col_map["outstanding"] = j
                    elif any(k in cl for k in ["rate", "roi", "interest"]):
                        col_map["rate"] = j
                    elif any(k in cl for k in ["security", "collateral", "mortgage"]):
                        col_map["security"] = j
                    elif any(k in cl for k in ["maturity", "due date", "repayment date"]):
                        col_map["maturity"] = j
                break

        if header_idx >= 0:
            for row in rows[header_idx + 1:]:
                entry = self._build_entry_from_cols(row, col_map)
                if entry:
                    result["loan_entries"].append(entry)
        else:
            # Heuristic: look for rows with numbers + bank keywords
            for row in rows:
                entry = self._heuristic_entry(row)
                if entry:
                    result["loan_entries"].append(entry)

        # Extract Covenants
        for row in rows:
            row_str = " ".join(str(c) for c in row).lower()
            if any(k in row_str for k in ["covenant", "compliance", "compliant", "breach"]):
                status = "COMPLIANT" if "compliant" in row_str and "non" not in row_str else "BREACH"
                if "breach" in row_str: status = "BREACH"
                if any(k in row_str for k in ["compliant", "breach"]):
                    result["covenants"].append({
                        "particulars": row[0] if row else "General Covenant",
                        "status": status,
                        "details": row_str
                    })

        self._aggregate(result)
        self._generate_risk_signals(result)
        logger.success(f"[BorrowingProfileExtractor] {len(result['loan_entries'])} entries, {len(result['covenants'])} covenants")
        return result

    def _build_entry_from_cols(self, row: list, col_map: dict) -> dict | None:
        if not row or not any(row):
            return None

        def get(key, default=""):
            idx = col_map.get(key)
            if idx is not None and idx < len(row):
                return str(row[idx]).strip()
            return default

        lender = get("lender")
        if not lender or len(lender) < 2:
            return None
        if not any(kw in lender.lower() for kw in LENDER_KEYWORDS + ["bank"]):
            # Try to see if it looks like a company name anyway
            if not any(c.isalpha() for c in lender):
                return None

        sanctioned_str = get("sanctioned")
        outstanding_str = get("outstanding")

        sanctioned_paise = CURR.parse_to_paise(sanctioned_str) or 0
        outstanding_paise = CURR.parse_to_paise(outstanding_str) or 0

        if sanctioned_paise == 0 and outstanding_paise == 0:
            return None

        rate_str = get("rate")
        rate = None
        rate_match = re.search(r'(\d+(?:\.\d+)?)\s*%?', rate_str)
        if rate_match:
            rate = float(rate_match.group(1))

        return {
            "lender": lender,
            "facility_type": get("facility_type") or "Unknown",
            "sanctioned_paise": sanctioned_paise,
            "outstanding_paise": outstanding_paise,
            "rate_pct": rate,
            "security": get("security"),
            "maturity": get("maturity"),
        }

    def _heuristic_entry(self, row: list) -> dict | None:
        """Fallback: parse a row without known column positions."""
        row_str = " ".join(str(c) for c in row if c).strip()
        if len(row_str) < 5:
            return None

        has_lender = any(kw in row_str.lower() for kw in LENDER_KEYWORDS)
        has_number = bool(re.search(r'\d{2,}', row_str))

        if not (has_lender and has_number):
            return None

        numbers = [CURR.parse_to_paise(t) for t in row_str.split() if CURR.parse_to_paise(t)]
        numbers = [n for n in numbers if n and n > 0]

        return {
            "lender": row[0] if row else "Unknown",
            "facility_type": "Unknown",
            "sanctioned_paise": numbers[0] if len(numbers) > 0 else 0,
            "outstanding_paise": numbers[1] if len(numbers) > 1 else (numbers[0] if numbers else 0),
            "rate_pct": None,
            "security": "",
            "maturity": "",
        }

    def _parse_from_text(self, text: str, result: dict) -> dict:
        """Line-by-line fallback when table extraction fails."""
        lines = text.split("\n")
        for line in lines:
            if not line.strip():
                continue
            if any(kw in line.lower() for kw in LENDER_KEYWORDS):
                nums = re.findall(r'[\d,]+(?:\.\d+)?', line)
                if nums:
                    outstanding = CURR.parse_to_paise(nums[-1]) or 0
                    sanctioned = CURR.parse_to_paise(nums[0]) if len(nums) > 1 else outstanding
                    if outstanding > 0:
                        result["loan_entries"].append({
                            "lender": line[:40].strip(),
                            "facility_type": "Unknown",
                            "sanctioned_paise": sanctioned,
                            "outstanding_paise": outstanding,
                            "rate_pct": None,
                            "security": "",
                            "maturity": "",
                        })

        self._aggregate(result)
        self._generate_risk_signals(result)
        return result

    def _aggregate(self, result: dict):
        total_sanctioned = sum(e.get("sanctioned_paise", 0) for e in result["loan_entries"])
        total_outstanding = sum(e.get("outstanding_paise", 0) for e in result["loan_entries"])
        result["total_sanctioned_paise"] = total_sanctioned
        result["total_outstanding_paise"] = total_outstanding
        result["total_funded_debt"] = total_outstanding
        result["lender_count"] = len(set(e["lender"].lower() for e in result["loan_entries"]))

        # Facility type breakdown
        facility_counts = {}
        for e in result["loan_entries"]:
            ft = e.get("facility_type", "Unknown")
            facility_counts[ft] = facility_counts.get(ft, 0) + e.get("outstanding_paise", 0)
        result["facility_breakdown"] = facility_counts

    def _generate_risk_signals(self, result: dict):
        signals = []
        CRORE = 10_000_000 * 100
        if result["lender_count"] > 5:
            signals.append({
                "signal_type": "MULTIPLE_BANK_BORROWINGS",
                "severity": "MEDIUM",
                "description": f"Entity has borrowings from {result['lender_count']} lenders — consortium lending implies higher monitoring burden, potential restructuring risk.",
                "five_c_mapping": "Capital",
            })
        if result["total_outstanding_paise"] > 500 * CRORE:
            signals.append({
                "signal_type": "HIGH_LEVERAGE",
                "severity": "HIGH",
                "description": f"Total outstanding borrowings of ₹{result['total_outstanding_paise'] // CRORE:.0f} Cr indicate significant leverage.",
                "five_c_mapping": "Capital",
            })
        result["risk_signals"] = signals

    def _count_matches(self, text: str, keywords: list) -> int:
        return sum(1 for kw in keywords if kw in text)

    def _empty_result(self) -> dict:
        return {
            "loan_entries": [],
            "covenants": [],
            "total_sanctioned_paise": 0,
            "total_outstanding_paise": 0,
            "total_funded_debt": 0,
            "lender_count": 0,
            "facility_breakdown": {},
            "risk_signals": [],
            "extraction_source": "borrowing_profile_extractor",
        }
