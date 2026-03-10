"""
Portfolio Performance Extractor (for NBFCs and financial entities)
Extracts: NPA%, collection efficiency, DPD bucket distribution, vintage analysis.
Handles PDF and XLSX formats.
"""
import re
from loguru import logger
from ingestion.normalization.currency_normalizer import CurrencyNormalizer

CURR = CurrencyNormalizer()
PCT_RE = re.compile(r'(\d{1,3}(?:\.\d{1,4})?)\s*%')
CRORE = 10_000_000 * 100


class PortfolioPerformanceExtractor:

    def extract(self, file_path: str) -> dict:
        result = self._empty_result()
        ext = file_path.lower().split(".")[-1]
        try:
            if ext in ("xlsx", "xls"):
                return self._extract_from_excel(file_path, result)
            else:
                return self._extract_from_pdf(file_path, result)
        except Exception as e:
            logger.warning(f"[PortfolioExtractor] Failed: {e}")
            return result

    # ── EXCEL ────────────────────────────────────────────────────────────────

    def _extract_from_excel(self, file_path: str, result: dict) -> dict:
        rows = []
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                for row in ws.iter_rows(values_only=True):
                    rows.append([str(c).strip() if c is not None else "" for c in row])
            wb.close()
        except Exception:
            try:
                import xlrd
                wb = xlrd.open_workbook(file_path)
                for sh in wb.sheets():
                    for r in range(sh.nrows):
                        rows.append([str(sh.cell_value(r, c)) for c in range(sh.ncols)])
            except Exception as e2:
                logger.warning(f"Portfolio Excel: {e2}")
                return result

        full_text = "\n".join(" ".join(row) for row in rows).lower()
        self._extract_from_text(full_text, result)
        self._extract_dpd_from_rows(rows, result)
        self._generate_risk_signals(result)
        return result

    # ── PDF ──────────────────────────────────────────────────────────────────

    def _extract_from_pdf(self, file_path: str, result: dict) -> dict:
        text = ""
        try:
            import fitz
            doc = fitz.open(file_path)
            for page in doc:
                text += page.get_text() + "\n"
            doc.close()
        except Exception:
            pass

        if len(text.strip()) < 100:
            try:
                import pdfplumber
                with pdfplumber.open(file_path) as pdf:
                    for page in pdf.pages:
                        t = page.extract_text()
                        if t:
                            text += t + "\n"
            except Exception:
                pass

        self._extract_from_text(text.lower(), result)

        # Try table extraction for DPD buckets
        try:
            import pdfplumber
            with pdfplumber.open(file_path) as pdf:
                for page in pdf.pages:
                    rows = page.extract_tables()
                    for table in rows:
                        if table:
                            self._extract_dpd_from_rows(table, result)
        except Exception:
            pass

        self._generate_risk_signals(result)
        return result

    # ── CORE EXTRACTION ──────────────────────────────────────────────────────

    def _extract_from_text(self, text: str, result: dict):
        """Extract KPIs from text using regex."""

        # Portfolio size / AUM
        for pattern in [
            r'(?:portfolio|aum|assets under management|loan book|total loan)\s*(?:size|outstanding)?\s*(?:of|:)?\s*([\d,]+(?:\.\d+)?)\s*(?:cr|crore|lakh)',
            r'total\s+(?:portfolio|loan)\s*(?:of|:)?\s*([\d,]+(?:\.\d+)?)\s*(?:cr|crore)',
        ]:
            m = re.search(pattern, text)
            if m:
                result["portfolio_size_paise"] = CURR.parse_to_paise(m.group(1) + " cr") or 0
                break

        # Gross NPA %
        for pattern in [
            r'gross\s+npa\s*(?:ratio|%|percentage)?\s*[:=]?\s*(\d{1,3}(?:\.\d{1,4})?)\s*%?',
            r'gnpa\s*[:=]?\s*(\d{1,3}(?:\.\d{1,4})?)\s*%?',
        ]:
            m = re.search(pattern, text)
            if m:
                v = float(m.group(1))
                if 0 <= v <= 100:
                    result["gross_npa_pct"] = v
                break

        # Net NPA %
        for pattern in [
            r'net\s+npa\s*(?:ratio|%|percentage)?\s*[:=]?\s*(\d{1,3}(?:\.\d{1,4})?)\s*%?',
            r'nnpa\s*[:=]?\s*(\d{1,3}(?:\.\d{1,4})?)\s*%?',
        ]:
            m = re.search(pattern, text)
            if m:
                v = float(m.group(1))
                if 0 <= v <= 100:
                    result["net_npa_pct"] = v
                break

        # Provision Coverage Ratio
        m = re.search(r'provision\s+coverage\s+(?:ratio|pcr)\s*[:=]?\s*(\d{1,3}(?:\.\d{1,4})?)\s*%?', text)
        if m:
            v = float(m.group(1))
            if 0 <= v <= 100:
                result["provision_coverage_pct"] = v

        # Collection Efficiency
        for pattern in [
            r'collection\s+efficiency\s*[:=]?\s*(\d{1,3}(?:\.\d{1,4})?)\s*%?',
            r'collection\s+rate\s*[:=]?\s*(\d{1,3}(?:\.\d{1,4})?)\s*%?',
        ]:
            m = re.search(pattern, text)
            if m:
                v = float(m.group(1))
                if 0 <= v <= 100:
                    result["collection_efficiency_pct"] = v
                break

        # CRAR / Capital Adequacy (for NBFCs)
        m = re.search(r'(?:crar|capital\s+adequacy\s+ratio|car)\s*[:=]?\s*(\d{1,3}(?:\.\d{1,4})?)\s*%?', text)
        if m:
            v = float(m.group(1))
            if 0 <= v <= 100:
                result["crar_pct"] = v

        # Write-offs
        m = re.search(r'write.?off\s*(?:amount|total)?\s*[:=]?\s*([\d,]+(?:\.\d+)?)\s*(?:cr|crore|lakh)', text)
        if m:
            result["write_off_paise"] = CURR.parse_to_paise(m.group(1) + " cr") or 0

    def _extract_dpd_from_rows(self, rows: list, result: dict):
        """Extract DPD bucket distribution from tables."""
        dpd_map = {
            "0_30": ["0-30", "0-29", "current", "0 to 30", "0 to 29", "std", "standard"],
            "30_60": ["30-60", "31-60", "31 to 60", "30 to 60", "dpd 30"],
            "60_90": ["60-90", "61-90", "61 to 90", "60 to 90", "dpd 60"],
            "90_plus": ["90+", "90 and above", "npa", ">90", "above 90", "substandard", "doubtful"],
        }

        for row in rows:
            if not row:
                continue
            label = str(row[0]).lower().strip() if row else ""
            for bucket_key, keywords in dpd_map.items():
                if any(kw in label for kw in keywords):
                    # Try to get the amount/percentage from subsequent columns
                    for cell in row[1:]:
                        cell_str = str(cell).replace(",", "").strip()
                        # Extract percentage
                        pct_m = re.search(r'(\d{1,3}(?:\.\d+)?)\s*%', cell_str)
                        if pct_m:
                            v = float(pct_m.group(1))
                            if 0 <= v <= 100:
                                result["dpd_buckets"][bucket_key] = v
                                break
                        # Extract raw amount
                        try:
                            v = float(cell_str)
                            if v > 0:
                                result["dpd_buckets"][bucket_key] = v
                                break
                        except ValueError:
                            continue

    def _generate_risk_signals(self, result: dict):
        signals = []

        gnpa = result.get("gross_npa_pct")
        if gnpa is not None:
            if gnpa > 10:
                signals.append({
                    "signal_type": "GROSS_NPA_CRITICAL",
                    "severity": "CRITICAL",
                    "description": f"Gross NPA ratio of {gnpa:.2f}% is above 10% — indicates severe asset quality deterioration. RBI guidelines flag this for Prompt Corrective Action (PCA).",
                    "five_c_mapping": "Capacity",
                })
            elif gnpa > 5:
                signals.append({
                    "signal_type": "GROSS_NPA_HIGH",
                    "severity": "HIGH",
                    "description": f"Gross NPA ratio of {gnpa:.2f}% exceeds 5% threshold — asset quality is stressed, recovery pipeline may be under pressure.",
                    "five_c_mapping": "Capacity",
                })

        ce = result.get("collection_efficiency_pct")
        if ce is not None and ce < 85:
            signals.append({
                "signal_type": "LOW_COLLECTION_EFFICIENCY",
                "severity": "HIGH" if ce < 75 else "MEDIUM",
                "description": f"Collection efficiency at {ce:.1f}% — below 85% indicates systemic collection failure, higher risk of NPA formation.",
                "five_c_mapping": "Capacity",
            })

        pcr = result.get("provision_coverage_pct")
        if pcr is not None and pcr < 70:
            signals.append({
                "signal_type": "LOW_PROVISION_COVERAGE",
                "severity": "MEDIUM",
                "description": f"Provision coverage ratio at {pcr:.1f}% is below RBI's recommended 70% — insufficient provisioning may mask true NPA impact.",
                "five_c_mapping": "Capital",
            })

        crar = result.get("crar_pct")
        if crar is not None and crar < 15:
            signals.append({
                "signal_type": "CRAR_BELOW_THRESHOLD",
                "severity": "HIGH" if crar < 10 else "MEDIUM",
                "description": f"Capital Adequacy (CRAR) at {crar:.1f}% is below the recommended 15% for NBFCs — capital buffer is thin for absorbing credit losses.",
                "five_c_mapping": "Capital",
            })

        result["risk_signals"] = signals
        logger.success(f"[PortfolioExtractor] GNPA={gnpa}%, CE={ce}%, PCR={pcr}%, signals={len(signals)}")

    def _empty_result(self) -> dict:
        return {
            "portfolio_size_paise": 0,
            "gross_npa_pct": None,
            "net_npa_pct": None,
            "provision_coverage_pct": None,
            "collection_efficiency_pct": None,
            "crar_pct": None,
            "write_off_paise": 0,
            "dpd_buckets": {
                "0_30": None,
                "30_60": None,
                "60_90": None,
                "90_plus": None,
            },
            "risk_signals": [],
            "extraction_source": "portfolio_performance_extractor",
        }
