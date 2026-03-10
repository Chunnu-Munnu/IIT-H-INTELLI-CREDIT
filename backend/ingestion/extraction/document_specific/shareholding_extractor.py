"""
Shareholding Pattern Extractor
Handles PDF and XLSX formats (BSE/NSE disclosure format + company annual report format).
Extracts: promoter %, FII/FPI %, DII %, public %, pledged shares %, total shareholders.
"""
import re
from loguru import logger
from ingestion.normalization.currency_normalizer import CurrencyNormalizer

CURR = CurrencyNormalizer()

# Regex patterns for percentage extraction
PCT_PATTERN = re.compile(r'(\d{1,3}(?:\.\d{1,4})?)\s*%?')
NUMBER_PATTERN = re.compile(r'[\d,]+(?:\.\d+)?')


class ShareholdingExtractor:

    def extract(self, file_path: str) -> dict:
        """Main entry. Returns structured shareholding dict."""
        result = self._empty_result()
        ext = file_path.lower().split(".")[-1]
        try:
            if ext in ("xlsx", "xls"):
                return self._extract_from_excel(file_path, result)
            else:
                return self._extract_from_pdf(file_path, result)
        except Exception as e:
            logger.warning(f"[ShareholdingExtractor] Failed: {e}")
            return result

    # ── EXCEL ────────────────────────────────────────────────────────────────

    def _extract_from_excel(self, file_path: str, result: dict) -> dict:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        except Exception:
            try:
                import xlrd
                return self._extract_xlrd(file_path, result)
            except Exception as e2:
                logger.warning(f"Excel open failed: {e2}")
                return result

        text_rows = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows(values_only=True):
                row_text = " ".join(str(c) if c is not None else "" for c in row).strip()
                if row_text:
                    text_rows.append(row_text)
        wb.close()

        full_text = "\n".join(text_rows).lower()
        return self._parse_text(full_text, result)

    def _extract_xlrd(self, file_path: str, result: dict) -> dict:
        import xlrd
        wb = xlrd.open_workbook(file_path)
        rows = []
        for sh in wb.sheets():
            for r in range(sh.nrows):
                row_text = " ".join(str(sh.cell_value(r, c)) for c in range(sh.ncols)).strip()
                if row_text:
                    rows.append(row_text)
        full_text = "\n".join(rows).lower()
        return self._parse_text(full_text, result)

    # ── PDF ──────────────────────────────────────────────────────────────────

    def _extract_from_pdf(self, file_path: str, result: dict) -> dict:
        text = ""
        try:
            import fitz
            doc = fitz.open(file_path)
            for page in doc:
                text += page.get_text().lower() + "\n"
            doc.close()
        except Exception as e:
            logger.warning(f"PDF text extract failed: {e}")

        if len(text.strip()) < 100:
            # Try pdfplumber
            try:
                import pdfplumber
                with pdfplumber.open(file_path) as pdf:
                    for page in pdf.pages:
                        t = page.extract_text()
                        if t:
                            text += t.lower() + "\n"
            except Exception as e2:
                logger.warning(f"pdfplumber fallback failed: {e2}")

        return self._parse_text(text, result)

    # ── PARSER ───────────────────────────────────────────────────────────────

    def _parse_text(self, text: str, result: dict) -> dict:
        lines = [l.strip() for l in text.split("\n") if l.strip()]

        # First pass for total shares and shareholders
        for line in lines:
            # Total shares (number)
            if self._matches(line, ["total number of shares", "total equity shares", "paid up shares"]):
                numbers = NUMBER_PATTERN.findall(line)
                for n in numbers:
                    try:
                        v = float(n.replace(",", ""))
                        if v > 1000: # Heuristic for share count
                            result["total_equity_shares"] = v
                            break
                    except: continue

            # Total shareholders
            if self._matches(line, ["total number of shareholders", "number of shareholders", "total shareholders", "total no. of shareholders"]):
                numbers = NUMBER_PATTERN.findall(line)
                for n in numbers:
                    try:
                        v = int(float(n.replace(",", "")))
                        if v > 10:
                            result["total_shareholders"] = v
                            break
                    except: continue

        # Second pass for percentages after getting total_equity_shares
        total_shares = result.get("total_equity_shares")
        for line in lines:
            val = self._extract_val_from_line(line, total_shares)
            if val is None: continue

            if self._matches(line, ["promoter", "promoters"]) and not self._matches(line, ["non-promoter", "non promoter"]):
                if result["promoter_pct"] is None: result["promoter_pct"] = val
            elif self._matches(line, ["fii", "fpi", "foreign institutional", "foreign portfolio"]):
                if result["fii_pct"] is None: result["fii_pct"] = val
            elif self._matches(line, ["dii", "domestic institutional", "mutual fund", "insurance"]):
                if result["dii_pct"] is None: result["dii_pct"] = val
            elif self._matches(line, ["public", "retail", "individual"]) and not self._matches(line, ["public limited"]):
                if result["public_pct"] is None: result["public_pct"] = val
            elif self._matches(line, ["pledged", "pledge", "encumbered"]):
                # Look for pledge specifically in this line
                if result["pledged_pct"] is None: result["pledged_pct"] = val

        # Compute public if missing
        known = [v for v in [result["promoter_pct"], result["fii_pct"], result["dii_pct"]] if v is not None]
        if result["public_pct"] is None and len(known) >= 2:
            result["public_pct"] = round(max(0, 100 - sum(known)), 2)

        # Risk flags
        result["risk_signals"] = []
        if result["promoter_pct"] is not None and result["promoter_pct"] < 51:
            result["risk_signals"].append({
                "signal_type": "LOW_PROMOTER_HOLDING",
                "severity": "HIGH",
                "description": f"Promoter holding {result['promoter_pct']}% is below 51% threshold — indicates weak promoter control, higher risk of hostile takeover or governance issues.",
                "five_c_mapping": "Character",
            })
        if result["pledged_pct"] is not None and result["pledged_pct"] > 30:
            result["risk_signals"].append({
                "signal_type": "HIGH_PLEDGED_SHARES",
                "severity": "CRITICAL" if result["pledged_pct"] > 60 else "HIGH",
                "description": f"{result['pledged_pct']}% of promoter shares are pledged — creates margin call risk if stock price falls, potentially forcing distress sale.",
                "five_c_mapping": "Character",
            })
        if result["fii_pct"] is not None and result["fii_pct"] > 40:
            result["risk_signals"].append({
                "signal_type": "HIGH_FII_EXPOSURE",
                "severity": "MEDIUM",
                "description": f"FII/FPI holding at {result['fii_pct']}% — sudden capital outflows could destabilize share price and collateral values.",
                "five_c_mapping": "Collateral",
            })

        result["extraction_source"] = "shareholding_extractor"
        logger.success(f"[ShareholdingExtractor] promoter={result['promoter_pct']}%, pledged={result['pledged_pct']}%, signals={len(result['risk_signals'])}")
        return result

    # ── Helpers ──────────────────────────────────────────────────────────────

    def _matches(self, line: str, keywords: list) -> bool:
        return any(kw in line.lower() for kw in keywords)

    def _extract_val_from_line(self, line: str, total_shares: float | None) -> float | None:
        """Extract a percentage value. If val > 100, treat as share count and divide by total_shares."""
        matches = NUMBER_PATTERN.findall(line)
        if not matches: return None
        
        # Take the last number in the line as it's usually the percentage or share count in these tables
        raw_val = float(matches[-1].replace(",", ""))
        
        if raw_val > 100:
            if total_shares and total_shares > 0:
                pct = (raw_val / total_shares) * 100
                return round(min(100.0, pct), 2)
            return None # Cannot convert count to pct without total
        
        return round(max(0.0, min(100.0, raw_val)), 2)

    def _empty_result(self) -> dict:
        return {
            "promoter_pct": None,
            "fii_pct": None,
            "dii_pct": None,
            "public_pct": None,
            "pledged_pct": None,
            "total_equity_shares": None,
            "total_shareholders": None,
            "risk_signals": [],
            "extraction_source": "shareholding_extractor",
        }
